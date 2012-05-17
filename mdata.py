import sys
import os
import re
import random
from collections import Counter
sys.path.append("openpyxl")
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
wb=load_workbook(filename=r"data_textmining.xlsx")
trainingf=open("training",'w');
testingf=open("testing",'w');
sheetfinal=wb.get_sheet_by_name(name='final')
death=0
alive=0
topk_token=5000
dic=[]
cnt = Counter()
deathf=[]
alivef=[]
finaltraindata=[]
for i in range(2,439):
        value=((sheetfinal.cell('E'+str(i)).value).lower()).strip()
        text=sheetfinal.cell('J'+str(i)).value+sheetfinal.cell('K'+str(i)).value+sheetfinal.cell('L'+str(i)).value
        text=text.strip()
        text=text.lower()
        if value=="yes":
           death=death+1
           deathf.append(text)
           if death<=175:
                for word in re.findall(r'\w+',text):
                    cnt[word]+=1
        elif value=="no":
             alive=alive+1
             alivef.append(text)
             if alive<=162:
                for word in re.findall(r'\w+',text):
                    cnt[word]+=1
        else:
             print "error"
             break
topk_token=cnt.most_common(topk_token)
feature=[[ "0" for col in range(len(topk_token))] for row in range(225+212)]
total=0
for item in topk_token:
       dic.append(item[0])
#print dic
#feature extraction
fileinter=0
for text in deathf:
    cnt=Counter()
    text=text.strip()
    text=text.lower()
    for word in re.findall(r'\w+',text):
        cnt[word]+=1
    wordinter=0
 
    for dicword in dic:
        if dicword in cnt:
           feature[fileinter][wordinter]="1"
        else:
             feature[fileinter][wordinter]="0"
        wordinter+=1
    fileinter+=1   
for i in range(0,225):
    if i<175:
             finaltraindata.append(str(i) + "\t1\t{" + ",".join(feature[i]) + "}\n");
             #trainingf.write(str(i) + "\t1\t{" + ",".join(feature[i]) + "}\n");
    else:
             testingf.write(str(i-175) + "\t1\t{" + ",".join(feature[i]) + "}\n");

for text in alivef:
    cnt=Counter()
    text=text.strip()
    text=text.lower()
    for word in re.findall(r'\w+',text):
        cnt[word]+=1
    wordinter=0
    for dicword in dic:
        if dicword in cnt:
           feature[fileinter][wordinter]="1"
        else:
             feature[fileinter][wordinter]="0"
        wordinter+=1
    fileinter+=1
for i in range(0,212):
    if i<162:
             finaltraindata.append(str(i+175) + "\t2\t{" + ",".join(feature[i+225]) + "}\n")
             #trainingf.write(str(i+175) + "\t2\t{" + ",".join(feature[i+225]) + "}\n");
    else:
             testingf.write(str(i-162+50) + "\t2\t{" + ",".join(feature[i+225]) + "}\n");
#write traning data into file
random.shuffle(finaltraindata)
for i in range(0,175+162):
    trainingf.write(finaltraindata[i]);
