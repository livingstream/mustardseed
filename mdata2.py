import sys
import os
import re
import random
from collections import Counter
sys.path.append("openpyxl")
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
wb=load_workbook(filename=r"data_textmining_daisy2.xlsx")
trainingf=open("training2dst",'w');
testingf=open("testing2dst",'w');
sheetfinal=wb.get_sheet_by_name(name='final')
death=0
alive=0
topk_token=9000
dic=[]
cnt = Counter()
deathf=[]
alivef=[]
traindata=[]
desTraindata=[]
for i in range(2,439):
        value=((sheetfinal.cell('E'+str(i)).value).lower()).strip()
        text=(sheetfinal.cell('S'+str(i)).value+sheetfinal.cell('T'+str(i)).value).strip().lower()
        if value=="yes":
           death=death+1
        elif value=="no":
             alive=alive+1
        if death<=175 and alive<=162:
           for word in re.findall(r'\w+',text):
               cnt[word]+=1
#get token features
topk_tokens=cnt.most_common(topk_token)
feature=[[ "0" for col in range(1+len(topk_tokens)+4)] for row in range(225+212)]
total=0
for item in topk_tokens:
       dic.append(item[0])
#print dic
featureInter=0
fileInter=0
for i in range(2,439):
        fileInter=fileInter+1
        featureInter=0
        value=((sheetfinal.cell('E'+str(i)).value).lower()).strip()
        text=(sheetfinal.cell('S'+str(i)).value+sheetfinal.cell('T'+str(i)).value).strip().lower()
        cnt=Counter()
        #class label
        if value=="yes":
             feature[fileInter-1][0]="1"
        elif value=="no":
             feature[fileInter-1][0]="2"
        #text token feature
        for word in re.findall(r'\w+',text):
            cnt[word]+=1
        for dicword in dic:
            featureInter=featureInter+1
            if dicword in cnt:
               feature[fileInter-1][featureInter]="1"
            else:
               feature[fileInter-1][featureInter]="0"
        #rifle feature, although I don't what it is
        featureInter=featureInter+1
        rifle=(str((sheetfinal.cell('F'+str(i)).value)).lower()).strip()
        feature[fileInter-1][featureInter]=rifle 
        #Gender feature
        featureInter=featureInter+1
        gender=(sheetfinal.cell('G'+str(i)).value.lower()).strip()
        if gender=="male":       
             feature[fileInter-1][featureInter]="1"
        elif gender=="female":
             feature[fileInter-1][featureInter]="0" 
        #age feature
        featureInter=featureInter+1
        age=(str(sheetfinal.cell('I'+str(i)).value).lower()).strip()
        feature[fileInter-1][featureInter]=age
        #cci feature
        featureInter=featureInter+1
        cci=(str(sheetfinal.cell('K'+str(i)).value).lower()).strip()
        feature[fileInter-1][featureInter]=cci

dead=0
traindead=0
testdead=0
alive=0 
trainalive=0
testalive=0

for i in range(0,437):
    if feature[i][0]=='1':
       dead=dead+1
       if dead<=175:
             traindead=traindead+1
             trainingf.write(str(traindead+trainalive) + "\t1\t{" + ",".join(feature[i][1:]) + "}\n");
       else: 
             testdead=testdead+1
             testingf.write(str(testdead+testalive) + "\t1\t{" + ",".join(feature[i][1:]) + "}\n");
    elif feature[i][0]=='2':
         alive=alive+1
         if alive<=162:
              trainalive=trainalive+1
              trainingf.write(str(traindead+trainalive) + "\t2\t{" + ",".join(feature[i][1:]) + "}\n")
         else:
              testalive=testalive+1
              testingf.write(str(testdead+testalive) + "\t2\t{" + ",".join(feature[i][1:]) + "}\n");
    else:
         print "error!"
         break
